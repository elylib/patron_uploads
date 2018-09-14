"""
We get Faculty/Staff data from payroll to load into Worldshare Management Services.

Currently this file comes from Melissa Cody.

This code reads that file and transforms the data into the format WMS expects.
It also constructs university IDs and let's the user know if any conflicts exist,
i.e. John Smith and Jane Smith both would have university IDs of jsmith, so the
user can check the university directory to clear up these situations.

This script allows for a less error-prone, more efficient, and more consistent method of loading
Fac/Staff accounts in WMS each year.
"""

import csv
from pprint import pprint
import string

import openpyxl

from constants import OCLC_FIELDS


def make_wsu_username(first, last):
    """
    Combine first and last names to make a WSU campus ID

    WSU uses the first letter of your first name and your last name to construct IDs.
    This gives us most of the names correctly, but you will have to manually check
    some against the directory because sometimes there is overlap (i.e. Steve and Suzanne Adams)
    and sometimes there is a space or punctuation in the last name that makes it reflect
    differently. Also, some emails will be wrong because some groups, like IT, have
    special email addresses. This has not traditionally been a problem, can revisit
    if it becomes one.
    """
    return first[0] + sanitize_last_name(last)


translator = str.maketrans('', '', string.punctuation)


def sanitize_last_name(last):
    """Get rid of spaces and punctuation to fit our typical user id rules"""
    out = last.translate(translator)
    return out.replace(' ', '')


def make_email(user):
    return '{0}@westfield.ma.edu'.format(user)


def main(file_location, today):
    year, month, day = today.year, today.month, today.day
    sheet = openpyxl.load_workbook(file_location).active
    with open(f'WEX_{year}_{month}_{day}_StaffPatrons_1_Tab_1.1.txt', 'w', newline="") as w:
        """
        This is not necessarily the ideal way to go about this (lining up two lists
        for field names and data rows), but this iteration is working. Some information
        that is constant among users is hardcoded.
        
        The WEX_2018_09_07_StaffPatrons_1_Tab_1.1.txt that is generated will need to be gone over to double check the rows with
        user ids in double_check that are printed out at the end. These are accounts that are
        likely to have errors because either a duplicate user id or a space in the last
        name.
        """
        writer = csv.writer(w, delimiter='\t')
        writer.writerow(OCLC_FIELDS)
        usernames = set()
        barcodes = set()
        double_check = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            row = [cell.value for cell in row]
            if row[6] in barcodes:
                # The file from payroll includes some people multiple times, i.e. if they teach in the
                # day and CGCE units, they are listed as employees in two areas. This will skip them
                # if we've already seen them.
                continue
            username = make_wsu_username(row[3].lower(), row[2].lower())
            # All these empty strings are to make sure the format of the tsv is just so
            # Should probably make the facstaff upload more like the student upload
            expiration_date = today.replace(year=today.year + 3).strftime('%Y-%m-%d')
            new_row = ['', row[3], '', row[2], '', '', '', '', '', '1410', row[6], username, 'urn:mace:oclc:idm:westfieldstatecollege:ldap',
                       'Faculty and Staff', '', expiration_date, '134347', row[4],
                       '577 Western Avenue', 'Westfield', 'MA', '01086', '', '', '', '', '', '', '', '', '',
                       make_email(username), '', '', '', '', '', '', '', '', '', '', '', '', '', '']
            writer.writerow(new_row)
            # This makes it so we can double check people who may be given an incorrect or duplicate
            # idAtSource value, so we can double check them in the directory
            if username in usernames or ' ' in row[2]:
                double_check.append(username)
            usernames.add(username)
            barcodes.add(row[6])
    pprint(double_check)
