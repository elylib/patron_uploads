import csv
from collections import namedtuple

import openpyxl

from constants import OCLC_FIELDS, BANNER_FIELDS, SCRUB_FIELDS


OCLCRecord = namedtuple('OCLCRecord', ' '.join(OCLC_FIELDS))
# Small metaprogramming black magic. Lets us initialize an OCLC Record with the Banner subset, and not worry about
# not immediately (or ever) setting other fields.
OCLCRecord.__new__.__defaults__ = (None,) * len(OCLCRecord._fields)


class Duplicate(Exception):
    def __str__(self):
        return "This student has already been seen."


seen = set()


def scrub_banner(banner):
    for key in banner.keys():
        if key in SCRUB_FIELDS:
            banner[key] = ''
    return banner


def banner_to_oclc(csv_row, today):
    """
    Move data from a Banner tsv row to an OCLCRecord.

    Make the tsv row from Banner by unpacking (*) the list of fields. Then convert that to a dict and unpack it (**)
    into an OCLCRecord namedtuple. This unpacking allows us to map the fields that already exist in the Banner report
    directly to the same-named fields in OCLCRecord. This is why we changed the name of circExpirationDate so this
    would map with errors.

    :param today: a datetime object so we can set expiration dates
    :param csv_row: Row of fields from the Banner tsv file.
    :return: An OCLCRecord namedtuple with the fields mapped from the Banner fields.
    :raise: Duplicate if we've already seen this one.
    """
    banner = scrub_banner(dict(zip(BANNER_FIELDS, csv_row)))

    # expiration date will be updated to 3 years from upload every time
    banner['oclcExpirationDate'] = today.replace(year=today.year + 3).strftime('%Y-%m-%d')

    if banner['idAtSource'] in seen or banner['barcode'] in seen:
        raise Duplicate

    seen.add(banner['idAtSource'])
    seen.add(banner['barcode'])
    return OCLCRecord(**banner)


def main(file_location, today):
    sheet = openpyxl.load_workbook(file_location).active
    with open('StudentPatrons_WEX_' + today.strftime('%Y%m%d') + '_1.txt', 'w', newline='') as w:
        writer = csv.writer(w, delimiter='\t')
        writer.writerow(OCLCRecord._fields)
        for patron_row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            try:
                patron = banner_to_oclc([cell.value for cell in patron_row], today)
            except Duplicate:
                # If there is a duplicate, you will need to look at it and use your judgement as to what
                # to do. OCLC will not accept duplicate patron names
                print([cell.value for cell in patron_row])
                continue

            writer.writerow(patron._asdict().values())
